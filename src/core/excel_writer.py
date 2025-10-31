"""
Écriture de données dans des fichiers Excel
"""
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelWriter:
    """
    Classe pour écrire des DataFrames dans un fichier Excel.
    """
    
    def __init__(self, output_path: Path, logger: Optional[logging.Logger] = None):
        """
        Initialiser l'écrivain Excel.
        
        Args:
            output_path: Chemin vers le fichier Excel de sortie
            logger: Logger optionnel
        """
        self.output_path = Path(output_path)
        self.logger = logger
        self.workbook: Optional[Workbook] = None
        
    def create_workbook(self) -> None:
        """
        Créer un nouveau classeur Excel.
        """
        self.workbook = Workbook()
        # Supprimer la feuille par défaut
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        if self.logger:
            self.logger.info("Nouveau classeur Excel créé")
    
    def add_dataframe(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        style_header: bool = True
    ) -> None:
        """
        Ajouter un DataFrame comme nouvelle feuille dans le classeur.
        
        Args:
            df: DataFrame à ajouter
            sheet_name: Nom de la feuille
            style_header: Appliquer un style à l'en-tête (gras, fond coloré)
        """
        if self.workbook is None:
            self.create_workbook()
        
        # Limiter la longueur du nom de la feuille (max 31 caractères pour Excel)
        sheet_name = sheet_name[:31]
        
        # Créer une nouvelle feuille
        ws = self.workbook.create_sheet(title=sheet_name)
        
        # Écrire les données du DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Styler la première ligne (en-têtes)
                if r_idx == 1 and style_header:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajuster automatiquement la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max 50 caractères
            ws.column_dimensions[column_letter].width = adjusted_width
        
        if self.logger:
            self.logger.info(
                f"Feuille '{sheet_name}' ajoutée: {len(df)} lignes, "
                f"{len(df.columns)} colonnes"
            )
    
    def add_multiple_dataframes(
        self,
        dataframes: Dict[str, pd.DataFrame],
        style_header: bool = True
    ) -> None:
        """
        Ajouter plusieurs DataFrames au classeur.
        
        Args:
            dataframes: Dictionnaire {nom_feuille: dataframe}
            style_header: Appliquer un style aux en-têtes
        """
        for sheet_name, df in dataframes.items():
            self.add_dataframe(df, sheet_name, style_header)
    
    def save(self) -> None:
        """
        Sauvegarder le classeur Excel sur le disque.
        
        Raises:
            Exception: Si le classeur ne peut pas être sauvegardé
        """
        if self.workbook is None:
            raise Exception("Aucun classeur à sauvegarder")
        
        try:
            self.workbook.save(self.output_path)
            
            if self.logger:
                self.logger.info(f"Fichier Excel sauvegardé: {self.output_path}")
        except Exception as e:
            if self.logger:
                self.logger.error(
                    f"Erreur lors de la sauvegarde du fichier: {str(e)}"
                )
            raise Exception(f"Impossible de sauvegarder le fichier Excel: {str(e)}")
    
    def get_file_size(self) -> tuple[int, str]:
        """
        Obtenir la taille du fichier Excel.
        
        Returns:
            Tuple (taille_en_octets, taille_formatée)
        """
        if not self.output_path.exists():
            return 0, "0 B"
        
        size_bytes = self.output_path.stat().st_size
        
        # Formater la taille
        if size_bytes < 1024:
            size_str = f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            size_str = f"{size_bytes / 1024:.2f} KB"
        elif size_bytes < 1024 * 1024 * 1024:
            size_str = f"{size_bytes / (1024 * 1024):.2f} MB"
        else:
            size_str = f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"
        
        return size_bytes, size_str
    
    def __enter__(self):
        """Support du context manager."""
        self.create_workbook()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Support du context manager."""
        if exc_type is None:
            self.save()
